#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
26年NC台账管理系统 - 云端部署版本
支持Render.com免费托管
"""
import os
import json
from datetime import datetime, timezone, timedelta
from flask import Flask, render_template, jsonify, request, send_file, Response
import gzip
import base64
import io
import hashlib
import re
import numpy as np
import pandas as pd

app = Flask(__name__)


# ========== 工具函数 ==========
def parse_amount(value):
    """智能解析金额：支持 ¥100, 1,000, 100元, 100.5 等各种格式，返回 float"""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value) if not pd.isna(value) else 0.0
    s = str(value).strip()
    if not s or s in ('nan', 'None', ''):
        return 0.0
    # 去除货币符号和文字
    s = re.sub(r'[¥$€£元%\s,，]', '', s)
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def get_pdf_font():
    """获取支持中文的 PDF 字体"""
    try:
        from fpdf import FPDF
        # 尝试使用系统自带的中文字体
        font_paths = [
            '/System/Library/Fonts/PingFang.ttc',
            '/System/Library/Fonts/STHeiti Light.ttc',
            '/System/Library/Fonts/Hiragino Sans GB.ttc',
            '/Library/Fonts/Arial Unicode.ttf',
        ]
        for fp in font_paths:
            if os.path.exists(fp):
                return fp
        return None  # fpdf2 内置不支持中文，需要添加字体
    except ImportError:
        return None


def render_analysis_html_report(df):
    """渲染精美HTML分析报告（支持浏览器打印为PDF），零依赖"""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import base64 as _b64, io as _io
    except ImportError:
        plt = None  # graceful fallback: no charts
        _b64 = None; _io = None

    # ── 全面防御：强制列类型正确（防止 float 传给 re 函数）───
    df = df.copy()  # 避免修改原数据
    if '日期' in df.columns:
        df['日期'] = df['日期'].apply(lambda x: '' if pd.isna(x) else str(x).strip())
    if '金额' in df.columns:
        df['金额'] = pd.to_numeric(df['金额'], errors='coerce').fillna(0)

    # ── 数据预处理：从日期字段派生年月信息 ──
    if '日期' in df.columns and '年月' not in df.columns:
        # 用 parse_date_flex 统一解析（与台账系统共用同一逻辑）
        df['年月'] = ''
        df['年份'] = 0
        df['月份'] = 0
        for idx in df.index:
            raw = df.loc[idx, '日期']
            parsed = parse_date_flex(raw)  # 与台账系统完全一致的解析逻辑
            if parsed and parsed not in ('', 'nan', 'nat'):
                try:
                    dt = pd.to_datetime(parsed)
                    df.loc[idx, '年月'] = dt.strftime('%Y-%m')
                    df.loc[idx, '年份'] = int(dt.year)
                    df.loc[idx, '月份'] = int(dt.month)
                except Exception:
                    # 兜底：直接从原字符串用正则提取年月
                    ds = str(raw) if not isinstance(raw, str) else raw
                    parts = re.split(r'[/\\-年]', ds)
                    if len(parts) >= 2:
                        try:
                            y, m = int(parts[0]), int(parts[1])
                            if 2000 <= y <= 2100 and 1 <= m <= 12:
                                df.loc[idx, '年月'] = f'{y:04d}-{m:02d}'
                                df.loc[idx, '年份'] = y
                                df.loc[idx, '月份'] = m
                        except Exception:
                            pass

    # ── 确保金额为数值型 ──
    if '金额' in df.columns:
        df['金额'] = pd.to_numeric(df['金额'], errors='coerce').fillna(0)

    # 中文字体（Mac + Linux）
    font_paths = [
        '/System/Library/Fonts/PingFang.ttc',
        '/System/Library/Fonts/STHeiti Light.ttc',
        '/System/Library/Fonts/Hiragino Sans GB.ttc',
        '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
        '/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf',
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
    ]
    chinese_font = next((f for f in font_paths if os.path.exists(f)), None)

    def make_chart(fig_size, func):
        if plt is None: return None
        try:
            plt.rcParams['font.family'] = chinese_font or 'DejaVu Sans'
            plt.rcParams['axes.unicode_minus'] = False
            fig, ax = func()
            buf = _io.BytesIO()
            plt.savefig(buf, format='png', dpi=150, bbox_inches='tight', facecolor='white')
            plt.close()
            buf.seek(0)
            return _b64.b64encode(buf.read()).decode()
        except Exception as e:
            print(f"[图表生成失败] {e}")
            return None

    # ── 图表1：月度破损金额对比 ──
    chart_monthly = None
    if '年月' in df.columns:
        valid_years = sorted([y for y in df['年份'].dropna().unique() if y > 2000])
        if valid_years:
            def monthly_func():
                fig, ax = plt.subplots(figsize=(12, 5))
                colors = ['#6366F1', '#F59E0B']
                x = list(range(1, 13))
                for idx, y in enumerate(valid_years[-2:]):
                    vals = [df[(df['年份']==y)&(df['月份']==m)]['金额'].sum()/10000 for m in range(1,13)]
                    ax.bar([i+(idx-0.5)*0.35 for i in x], vals, 0.32,
                           label=f'{int(y)}年', color=colors[idx], alpha=0.85)
                ax.set_xlabel('月份', fontsize=11)
                ax.set_ylabel('破损金额（万元）', fontsize=11)
                ax.set_title('月度破损金额对比', fontsize=14, fontweight='bold')
                ax.set_xticks(x); ax.set_xticklabels([f'{m}月' for m in x])
                ax.legend(); ax.grid(axis='y', alpha=0.2)
                ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
                return fig, ax
            try: chart_monthly = make_chart((12,5), monthly_func)
            except Exception as e: print(f"图表monthly失败: {e}")

    # ── 图表2：破损金额趋势折线图 ──
    chart_trend = None
    if '年月' in df.columns:
        def trend_func():
            fig, ax = plt.subplots(figsize=(12, 4.5))
            monthly = df.groupby('年月')['金额'].sum().reset_index().sort_values('年月')
            x, y = list(range(len(monthly))), monthly['金额'].values / 10000
            ax.plot(x, y, marker='o', linewidth=2.5, markersize=7, color='#6366F1')
            ax.fill_between(x, y, alpha=0.1, color='#6366F1')
            mx = y.argmax()
            ax.annotate(f'最高: {y[mx]:.2f}万', xy=(x[mx], y[mx]),
                        xytext=(max(0,x[mx]-2), y[mx]+0.3),
                        arrowprops=dict(arrowstyle='->', color='#EF4444'),
                        fontsize=10, color='#EF4444')
            ax.set_xlabel('月份', fontsize=11); ax.set_ylabel('破损金额（万元）', fontsize=11)
            ax.set_title('破损金额趋势', fontsize=14, fontweight='bold')
            ax.set_xticks(x); ax.set_xticklabels(monthly['年月'].values, rotation=45, ha='right', fontsize=9)
            ax.grid(alpha=0.2); ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
            return fig, ax
        try: chart_trend = make_chart((12,4.5), trend_func)
        except Exception as e: print(f"图表trend失败: {e}")

    # ── 图表3：责任方饼图 ──
    chart_pie = None
    if '责任方' in df.columns:
        resp = df.groupby('责任方')['金额'].sum().sort_values(ascending=False).head(8)
        if len(resp) > 0:
            def pie_func():
                fig, ax = plt.subplots(figsize=(8, 6))
                colors = ['#6366F1','#F59E0B','#10B981','#EF4444','#8B5CF6','#EC4899','#06B6D4','#84CC16']
                wedges, texts, autotexts = ax.pie(resp.values, labels=resp.index, autopct='%1.1f%%',
                                                  startangle=90, colors=colors[:len(resp)], pctdistance=0.75)
                for at in autotexts: at.set_fontsize(9); at.set_color('white')
                for t in texts: t.set_fontsize(9)
                ax.set_title('责任方破损金额占比', fontsize=13, fontweight='bold')
                return fig, ax
            try: chart_pie = make_chart((8,6), pie_func)
            except Exception as e: print(f"图表pie失败: {e}")

    # ── 图表4：商品详情柱状图 ──
    chart_product = None
    if '商品详情' in df.columns:
        prod = df.groupby('商品详情')['金额'].sum().sort_values(ascending=False).head(10)
        if len(prod) > 0:
            def prod_func():
                fig, ax = plt.subplots(figsize=(10, 5))
                colors = ['#6366F1','#8B5CF6','#A78BFA','#C4B5FD','#DDD6FE',
                          '#F59E0B','#FBBF24','#FCD34D','#FDE68A','#FEF3C7']
                bars = ax.barh(range(len(prod)), prod.values / 10000, color=colors[:len(prod)], height=0.6)
                ax.set_yticks(range(len(prod)))
                ax.set_yticklabels([str(p)[:20] for p in prod.index], fontsize=9)
                ax.set_xlabel('破损金额（万元）', fontsize=11)
                ax.set_title('商品破损金额 TOP10', fontsize=13, fontweight='bold')
                for i, (v, b) in enumerate(zip(prod.values / 10000, bars)):
                    ax.text(v + 0.1, b.get_y() + b.get_height()/2, f'{v:.2f}万',
                            va='center', fontsize=9, color='#374151')
                ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
                ax.invert_yaxis()
                ax.grid(axis='x', alpha=0.2)
                return fig, ax
            try: chart_product = make_chart((10,5), prod_func)
            except Exception as e: print(f"图表product失败: {e}")

    # ── 统计数据 ──
    total_records = len(df)
    total_amount = df['金额'].sum()
    valid_amount_records = (df['金额'] > 0).sum()
    valid_years = sorted([int(y) for y in df['年份'].dropna().unique() if y > 2000])

    # 月度环比
    monthly_rows = []
    if '年月' in df.columns:
        grp = df.groupby('年月')['金额'].sum().reset_index().sort_values('年月').tail(12)
        prev_val = None
        for _, row in grp.iterrows():
            amt = row['金额']
            if prev_val and prev_val != 0:
                chg = (amt - prev_val) / prev_val * 100
                icon = '🔴↑' if chg > 0 else ('🟢↓' if chg < 0 else '⚪-')
                chg_color = "#ef4444" if chg > 0 else "#10b981"
                chg_str = f'<span style="color:{chg_color}">{chg:+.1f}%</span>'
            else:
                chg_str = '-'; icon = ''
            monthly_rows.append({'年月': str(row['年月']), '金额': amt, '变化': chg_str, '图标': icon})
            prev_val = amt

    # 年同比
    yoy_html = ''
    if len(valid_years) >= 2:
        y_c, y_p = valid_years[-1], valid_years[-2]
        a_c = df[df['年份']==y_c]['金额'].sum()
        a_p = df[df['年份']==y_p]['金额'].sum()
        c_c = len(df[df['年份']==y_c])
        c_p = len(df[df['年份']==y_p])
        amt_chg = (a_c - a_p) / a_p * 100 if a_p > 0 else 0
        clr = '#ef4444' if amt_chg > 0 else '#10b981'
        yoy_html = f"<div class='yoy-block'>" \
            f"<div class='yoy-item'><div class='yoy-label'>{y_p}年</div><div class='yoy-value' style='color:#6366F1'>¥{a_p:,.0f}</div><div class='yoy-count'>{c_p}条</div></div>" \
            f"<div class='yoy-arrow'>→</div>" \
            f"<div class='yoy-item'><div class='yoy-label'>{y_c}年</div><div class='yoy-value' style='color:#ef4444'>¥{a_c:,.0f}</div><div class='yoy-count'>{c_c}条</div></div>" \
            f"<div class='yoy-chg'><span style='color:{clr}'>{'↑' if amt_chg>0 else '↓'} {abs(amt_chg):.1f}%</span></div></div>"

    # 月度环比表格行
    mid = len(monthly_rows) // 2
    tbody = ''
    for i in range(mid):
        r1 = monthly_rows[i]
        r2 = monthly_rows[i+mid] if i+mid < len(monthly_rows) else None
        tbody += f"<tr><td>{r1['年月']}</td><td>¥{r1['金额']:,.0f}</td><td>{r1['变化']}</td><td>{r1['图标']}</td>"
        if r2:
            tbody += f"<td>{r2['年月']}</td><td>¥{r2['金额']:,.0f}</td><td>{r2['变化']}</td><td>{r2['图标']}</td></tr>"
        else:
            tbody += "<td>-</td><td>-</td><td>-</td><td>-</td></tr>"

    # 责任方详细
    resp_rows = ''
    if '责任方' in df.columns:
        rd = df.groupby('责任方').agg(次数=('金额','count'), 金额=('金额','sum')).reset_index().sort_values('金额', ascending=False)
        for _, r in rd.iterrows():
            pct = r['金额']/total_amount*100 if total_amount>0 else 0
            resp_rows += f"<tr><td style='text-align:left;font-weight:600'>{str(r['责任方'])[:18]}</td>" \
                f"<td>{int(r['次数'])}</td><td>¥{r['金额']:,.0f}</td>" \
                f"<td><div style='font-weight:700;color:#5b21b6'>{pct:.1f}%</div>" \
                f"<div class='bar'><div class='fill' style='width:{min(pct,100):.1f}%'></div></div></td>" \
                f"<td>{'🔴' if pct>20 else ('🟡' if pct>5 else '🟢')}</td></tr>"


    # ── 逐月对比明细表（同比分析）──
    month_compare_html = ''
    if len(valid_years) >= 1 and '年月' in df.columns:
        month_compare_html = '<div class="section page-break"><div class="section-title"><span>📊</span>逐月对比明细（同比分析）</div>' \
            '<table class="month-compare-table"><thead><tr><th>月份</th>'
        for y in valid_years:
            month_compare_html += f'<th>{int(y)}年金额</th><th>同比</th>'
        month_compare_html += '</tr></thead><tbody>'
        for m in range(1, 13):
            month_compare_html += f'<tr><td>{m}月</td>'
            prev_amt = None
            for y in valid_years:
                amt = df[(df['年份']==y)&(df['月份']==m)]['金额'].sum()
                if prev_amt is not None and prev_amt > 0 and amt > 0:
                    chg = (amt - prev_amt) / prev_amt * 100
                    chg_str = f'<span style="color:{"#ef4444" if chg>0 else "#10b981"}">{chg:+.1f}%</span>'
                else:
                    chg_str = '-'
                month_compare_html += f'<td>¥{amt:,.0f}</td><td>{chg_str}</td>'
                prev_amt = amt
            month_compare_html += '</tr>'
        month_compare_html += '</tbody></table></div>'

    # ── 数据洞察与结论 ──
    insights_html = '<div class="section page-break"><div class="section-title"><span>💡</span>数据洞察与结论</div><div class="insights-box">'
    # 最高/最低月份
    if '年月' in df.columns and len(df) > 0:
        try:
            monthly_amt = df.groupby('年月')['金额'].sum().reset_index().sort_values('年月')
            if len(monthly_amt) > 0:
                row_max = monthly_amt.loc[monthly_amt['金额'].idxmax()]
                row_min = monthly_amt.loc[monthly_amt['金额'].idxmin()]
                insights_html += f'<div class="insight-item">📌 <strong>破损金额最高的月份：</strong>{row_max["年月"]}（¥{row_max["金额"]:,.0f}）</div>'
                if row_min["年月"] != row_max["年月"]:
                    insights_html += f'<div class="insight-item">📌 <strong>破损金额最低的月份：</strong>{row_min["年月"]}（¥{row_min["金额"]:,.0f}）</div>'
        except Exception as e:
            print(f"洞察-月份分析失败: {e}")
    # 责任方分析
    if '责任方' in df.columns:
        try:
            resp = df.groupby('责任方')['金额'].sum().sort_values(ascending=False)
            if len(resp) > 0:
                top_resp = resp.index[0]
                top_pct = resp.iloc[0] / total_amount * 100 if total_amount > 0 else 0
                insights_html += f'<div class="insight-item">🏢 <strong>主要破损责任方：</strong>{top_resp}（占比{top_pct:.1f}%）——建议重点跟进！</div>'
                if len(resp) >= 2:
                    insights_html += f'<div class="insight-item">🏢 <strong>次要破损责任方：</strong>{resp.index[1]}（占比{resp.iloc[1]/total_amount*100:.1f}%）</div>'
        except Exception as e:
            print(f"洞察-责任方分析失败: {e}")
    # 趋势预警
    if '年月' in df.columns:
        try:
            monthly_sorted = df.groupby('年月')['金额'].sum().reset_index().sort_values('年月')
            if len(monthly_sorted) >= 3:
                last3 = monthly_sorted.tail(3)['金额'].values
                if last3[-1] > last3[-2] and last3[-2] >= last3[-3]:
                    insights_html += '<div class="insight-item">⚠️ <strong>趋势预警：</strong>破损金额连续上升，请立即排查原因！</div>'
                elif last3[-1] < last3[-2] and last3[-2] <= last3[-3]:
                    insights_html += '<div class="insight-item">✅ <strong>趋势向好：</strong>破损金额持续下降，管理措施有效！</div>'
        except Exception as e:
            print(f"洞察-趋势分析失败: {e}")
    insights_html += '<div class="insight-item">📝 <strong>管理建议：</strong>建议每月复盘破损数据，对高频破损责任方和产品建立预警机制，及时采取改进措施。</div>'
    insights_html += '</div></div>'
    years_str = ', '.join(f'{y}年' for y in valid_years) if valid_years else '-'
    chart_m_img = f'<img src="data:image/png;base64,{chart_monthly}" alt="月度对比">' if chart_monthly else '<p style="color:#9ca3af;padding:40px">数据不足</p>'
    chart_t_img = f'<img src="data:image/png;base64,{chart_trend}" alt="趋势">' if chart_trend else '<p style="color:#9ca3af;padding:40px">数据不足</p>'
    chart_p_img = f'<img src="data:image/png;base64,{chart_pie}" alt="饼图">' if chart_pie else '<p style="color:#9ca3af;padding:40px">数据不足</p>'
    chart_p_img_prod = f'<img src="data:image/png;base64,{chart_product}" alt="商品排行">' if chart_product else '<p style="color:#9ca3af;padding:40px">数据不足</p>'

    return f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>NC台账数据分析报告</title>
<style>
@page{{size:A4;margin:15mm 15mm 20mm}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,"PingFang SC","Microsoft YaHei",sans-serif;background:#f8f7ff;color:#1e1b4b;font-size:13px;padding:20px}}
.page{{background:white;border-radius:16px;padding:32px;max-width:210mm;margin:0 auto;box-shadow:0 4px 24px rgba(120,84,244,0.1)}}
.report-header{{text-align:center;margin-bottom:28px;padding-bottom:20px;border-bottom:2px solid #ede9fe}}
.report-title{{font-size:26px;font-weight:800;color:#5b21b6;margin-bottom:6px}}
.report-meta{{color:#6b7280;font-size:12px;margin-top:6px}}
.kpi-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:28px}}
.kpi-card{{background:linear-gradient(135deg,#f5f3ff,#ede9fe);border-radius:14px;padding:18px 14px;text-align:center;border:1px solid #ddd6fe}}
.kpi-value{{font-size:22px;font-weight:800;color:#5b21b6}}
.kpi-sub{{color:#6b7280;font-size:11px;margin-top:4px}}
.yoy-block{{display:flex;align-items:center;justify-content:center;gap:16px;margin:20px 0;padding:16px;background:#faf9ff;border-radius:12px;border:1px solid #e9e3ff}}
.yoy-item{{text-align:center}}
.yoy-label{{font-size:12px;color:#6b7280;margin-bottom:4px}}
.yoy-value{{font-size:18px;font-weight:700}}
.yoy-count{{font-size:11px;color:#9ca3af}}
.yoy-arrow{{font-size:22px;color:#a78bfa}}
.yoy-chg{{font-size:18px;font-weight:800;padding:8px 14px;background:white;border-radius:10px;border:1px solid #e9e3ff}}
.section{{margin-bottom:28px}}
.section-title{{font-size:16px;font-weight:700;color:#1e1b4b;margin-bottom:14px;padding-bottom:8px;border-bottom:2px solid #ede9fe;display:flex;align-items:center;gap:8px}}
.chart-wrap{{background:#faf9ff;border-radius:12px;padding:16px;border:1px solid #ede9fe;text-align:center}}
.chart-wrap img{{max-width:100%;height:auto;border-radius:8px}}
.charts-row{{display:grid;grid-template-columns:1fr 1fr;gap:16px}}
.monthly-table{{width:100%;border-collapse:collapse;font-size:12px}}
.monthly-table th{{background:#5b21b6;color:white;padding:10px 12px;text-align:center;border-radius:8px}}
.monthly-table td{{padding:9px 12px;text-align:center;border-bottom:1px solid #f0eeff}}
.monthly-table tr:nth-child(even) td{{background:#faf9ff}}
.resp-table{{width:100%;border-collapse:collapse;font-size:12px}}
.resp-table th{{background:linear-gradient(135deg,#5b21b6,#7c3aed);color:white;padding:10px 14px;text-align:center}}
.resp-table td{{padding:9px 14px;text-align:center;border-bottom:1px solid #f0eeff}}
.resp-table tr:nth-child(even) td{{background:#faf9ff}}
.bar{{height:6px;background:#ede9fe;border-radius:3px;margin-top:4px}}
.fill{{height:100%;background:linear-gradient(90deg,#6366F1,#a78bfa);border-radius:3px;transition:width .3s}}
  .insights-box{{background:#faf9ff;border-radius:12px;padding:20px;border:1px solid #e9e3ff;margin-top:12px}}
  .insight-item{{padding:10px 0;border-bottom:1px solid #f0eeff;font-size:13px;line-height:1.6}}
  .insight-item:last-child{{border-bottom:none}}
  .month-compare-table{{width:100%;border-collapse:collapse;font-size:12px}}
  .month-compare-table th{{background:#5b21b6;color:white;padding:8px 10px;text-align:center}}
  .month-compare-table td{{padding:8px 10px;text-align:center;border-bottom:1px solid #f0eeff}}
  .month-compare-table tr:nth-child(even) td{{background:#faf9ff}}
.footer{{text-align:center;padding-top:16px;margin-top:20px;border-top:1px solid #ede9fe;color:#9ca3af;font-size:11px}}
.print-note{{background:#fef3c7;border-radius:8px;padding:10px 16px;margin-bottom:20px;font-size:12px;color:#92400e;text-align:center}}
.no-print{{display:block}}
@media print{{body{{background:white;padding:0}}.page{{box-shadow:none;border-radius:0;padding:20px}}.no-print{{display:none!important}}.page-break{{page-break-after:always}}}}
</style>
</head>
<body>
<div class="page">
<div class="print-note no-print">💡 <strong>打印PDF：</strong>按 <kbd>Ctrl+P</kbd>（Mac: <kbd>⌘P</kbd>）→ 目标选"保存为PDF" → 保存</div>
<div class="report-header">
  <div class="report-title">📊 NC台账数据分析报告</div>
  <div class="report-meta">生成时间：{datetime.now(timezone(timedelta(hours=8))).strftime('%Y-%m-%d %H:%M')} &nbsp;|&nbsp; NC台账管理系统</div>
</div>
<div class="kpi-grid">
  <div class="kpi-card"><div class="kpi-value">{total_records}</div><div class="kpi-sub">总记录数</div></div>
  <div class="kpi-card"><div class="kpi-value">¥{total_amount:,.0f}</div><div class="kpi-sub">总破损金额</div></div>
  <div class="kpi-card"><div class="kpi-value">{valid_amount_records}</div><div class="kpi-sub">有金额记录</div></div>
  <div class="kpi-card"><div class="kpi-value">{years_str}</div><div class="kpi-sub">覆盖年份</div></div>
</div>
{yoy_html}
<div class="section page-break">
  <div class="section-title"><span>📈</span>月度破损金额对比</div>
  <div class="chart-wrap">{chart_m_img}</div>
</div>
<div class="section">
  <div class="charts-row">
    <div>
      <div class="section-title"><span>📉</span>破损金额趋势</div>
      <div class="chart-wrap">{chart_t_img}</div>
    </div>
    <div>
      <div class="section-title"><span>🥧</span>责任方占比</div>
      <div class="chart-wrap">{chart_p_img}</div>
    </div>
  </div>
</div>
<div class="section page-break">
  <div class="section-title"><span>📅</span>月度环比分析（🔴↑上涨 🟢↓下降）</div>
  <table class="monthly-table">
    <thead><tr><th>月份</th><th>破损金额</th><th>环比变化</th><th>趋势</th><th>月份</th><th>破损金额</th><th>环比变化</th><th>趋势</th></tr></thead>
    <tbody>{tbody}</tbody>
  </table>
</div>
<div class="section">
  <div class="section-title"><span>🏢</span>责任方详细数据</div>
  <table class="resp-table">
    <thead><tr><th>责任方</th><th>破损次数</th><th>破损金额</th><th>金额占比</th><th>等级</th></tr></thead>
    <tbody>{resp_rows}</tbody>
  </table>
</div>
{month_compare_html}
{insights_html}
<div class="footer">NC台账管理系统 · 数据分析报告 · Generated by QClaw</div>
</div>
</body>
</html>'''


# Gzip压缩 - 只压缩文本内容
@app.after_request
def compress_response(response):
    accept_encoding = request.headers.get('Accept-Encoding', '').lower()
    if 'gzip' not in accept_encoding or response.status_code != 200:
        return response
    # 排除二进制文件（Excel、图片等）
    content_type = response.content_type or ''
    if any(x in content_type for x in ['octet-stream', 'excel', 'spreadsheet', 'image', 'pdf']):
        return response
    if 'attachment' in response.headers.get('Content-Disposition', ''):
        return response
    try:
        content = response.get_data()
        gzip_response = Response(gzip.compress(content), content_type=content_type)
        gzip_response.headers['Content-Encoding'] = 'gzip'
        gzip_response.headers['Vary'] = 'Accept-Encoding'
        return gzip_response
    except:
        return response

# Render.com环境变量或本地文件
EXCEL_FILE = os.environ.get('EXCEL_FILE', os.path.expanduser('~/Desktop/26年NC台账勿删.xlsx'))
DATA_FILE = '/tmp/nc_account_data.json'
EMBEDDED_DATA_FILE = os.path.join(os.path.dirname(__file__), 'embedded_data.json')

def load_data():
    """加载数据 - 优先临时文件（最新修改），其次嵌入文件"""
    # 1. 优先从临时文件加载（运行时修改的数据，最新）
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if data and len(data) > 0:
                    return data
        except:
            pass
    
    # 2. 其次从嵌入的数据文件加载（GitHub同步的数据）
    if os.path.exists(EMBEDDED_DATA_FILE):
        try:
            with open(EMBEDDED_DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if data and len(data) > 0:
                    # 复制到临时文件
                    save_data(data)
                    return data
        except:
            pass
    
    return []



# ==================== 操作日志功能 ====================

LOG_FILE = '/tmp/operation_logs.json'
MAX_LOGS = 500  # 最多保留500条日志

def add_log(action, detail, user='system'):
    """添加操作日志"""
    try:
        logs = []
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r') as f:
                logs = json.load(f)
        
        log_entry = {
            'time': datetime.now(timezone.utc).isoformat(),
            'action': action,
            'detail': detail,
            'user': user,
            'ip': request.remote_addr if request else 'local'
        }
        
        logs.insert(0, log_entry)
        
        # 限制日志数量
        if len(logs) > MAX_LOGS:
            logs = logs[:MAX_LOGS]
        
        with open(LOG_FILE, 'w') as f:
            json.dump(logs, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f'日志记录失败: {e}')

def get_logs(limit=100):
    """获取操作日志"""
    try:
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r') as f:
                logs = json.load(f)
                return logs[:limit]
    except:
        pass
    return []

@app.route('/api/logs')
def api_logs():
    """获取操作日志API"""
    limit = request.args.get('limit', 100, type=int)
    logs = get_logs(limit)
    return jsonify({'success': True, 'logs': logs})

@app.route('/api/logs/export')
def export_logs():
    """导出操作日志"""
    try:
        logs = get_logs(MAX_LOGS)
        # 生成CSV内容
        csv_content = '时间,操作,详情,用户,IP\n'
        for log in logs:
            time_str = log.get('time', '').replace('T', ' ').split('.')[0]
            action = log.get('action', '')
            detail = log.get('detail', '').replace('"', '""')
            user = log.get('user', 'system')
            ip = log.get('ip', '')
            csv_content += f'"{time_str}","{action}","{detail}","{user}","{ip}"\n'
        
        # 生成文件名
        filename = f'操作日志_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")}.csv'
        return send_file(
            io.BytesIO(csv_content.encode('utf-8-sig')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def save_data(data):
    """保存数据到临时文件"""
    try:
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except:
        pass
    # 同时尝试保存到嵌入文件（本地环境）
    try:
        with open(EMBEDDED_DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except:
        pass

def sync_to_github(data):
    """同步数据到GitHub（需要配置GITHUB_TOKEN环境变量）"""
    token = os.environ.get('GITHUB_TOKEN')
    if not token:
        return False, '未配置GITHUB_TOKEN'
    
    try:
        import requests
        
        # GitHub API 配置
        owner = 'NC0319'
        repo = 'nc-account-system'
        path = 'embedded_data.json'
        
        # 获取当前文件的 SHA
        api_url = f'https://api.github.com/repos/{owner}/{repo}/contents/{path}'
        headers = {
            'Authorization': f'token {token}',
            'Accept': 'application/vnd.github.v3+json'
        }
        
        # 添加10秒超时，防止卡死
        r = requests.get(api_url, headers=headers, timeout=10)
        sha = r.json().get('sha', '') if r.status_code == 200 else ''
        
        # 准备新内容
        content = json.dumps(data, ensure_ascii=False, indent=2)
        content_b64 = base64.b64encode(content.encode('utf-8')).decode('utf-8')
        
        # 更新文件
        payload = {
            'message': '自动同步数据更新',
            'content': content_b64,
            'sha': sha
        }
        
        r = requests.put(api_url, headers=headers, json=payload, timeout=10)
        if r.status_code in [200, 201]:
            return True, ''
        else:
            return False, f'GitHub API返回错误: {r.status_code} {r.text[:200]}'
    except Exception as e:
        err_msg = str(e)
        print(f"同步GitHub失败: {err_msg}")
        return False, err_msg

def save_to_excel(data):
    """同步保存到Excel"""
    try:
        df = pd.DataFrame(data)
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        
        # 云端写入到挂载的磁盘，本地写入桌面
        excel_path = os.environ.get('EXCEL_FILE', EXCEL_FILE)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='台账')
        return True
    except Exception as e:
        print(f"保存Excel失败: {e}")
        return False

@app.route('/health')
def health():
    return 'OK', 200

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/data', methods=['GET'])
def get_data():
    return jsonify(load_data())

@app.route('/api/data', methods=['POST'])
def add_data():
    data = load_data()
    data.append(request.json)
    save_data(data)
    sync_to_github(data)
    return jsonify({'success': True})

@app.route('/api/data/<int:index>', methods=['PUT', 'PATCH'])
def update_data(index):
    data = load_data()
    if 0 <= index < len(data):
        # PATCH: 部分更新（合并）；PUT: 整条替换
        if request.method == 'PATCH':
            data[index].update(request.json)
        else:
            data[index] = request.json
        add_log('更新数据', f'索引{index}, 包裹号: {data[index].get("包裹号", "")}')
        save_data(data)
        sync_to_github(data)
        return jsonify({'success': True})
    return jsonify({'success': False}), 404

@app.route('/api/data/<int:index>', methods=['DELETE'])
def delete_data(index):
    data = load_data()
    if 0 <= index < len(data):
        pkg = data[index].get('包裹号', '')
        data.pop(index)
        add_log('删除数据', f'包裹号: {pkg}')
        save_data(data)
        sync_to_github(data)
        return jsonify({'success': True})
    return jsonify({'success': False}), 404

@app.route('/api/export')
def export_excel():
    data = load_data()
    excel_path = '/tmp/nc_account_export.xlsx'
    df = pd.DataFrame(data)
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    df.to_excel(excel_path, index=False)
    return send_file(excel_path, as_attachment=True)

def standardize_columns(df):
    """标准化列名，映射常见变体到标准字段名（含智能关键词匹配）"""
    column_mapping = {
        # 凭证相关
        '凭证号': '凭证', '单据号': '凭证', '凭证编号': '凭证', '单号': '凭证',
        '订单号': '凭证', '运单号': '凭证', '快递单号': '凭证',
        # 日期相关（精确 + 常见业务变体）
        '日期': '日期', '时间': '日期', '日期时间': '日期', '发生时间': '日期', '发生日期': '日期',
        '生产日期': '日期', '下单日期': '日期', '到货日期': '日期', '收货日期': '日期',
        '出库日期': '日期', '入库日期': '日期', '开单日期': '日期', '记账日期': '日期',
        '业务日期': '日期', '交易日期': '日期', '单据日期': '日期', '订单日期': '日期',
        '发货日期': '日期', '处理日期': '日期', '登记日期': '日期', '录入日期': '日期',
        '上报日期': '日期', '创建日期': '日期', '生成日期': '日期', '日期1': '日期',
        '开班日期': '日期',
        # 班次相关
        '班别': '班次', '班制': '班次', '班组': '班次', '白夜班': '班次', '早晚班': '班次',
        '班次类型': '班次',
        # 路由相关
        '路由状态': '路由', '物流状态': '路由', '物流': '路由', '物流信息': '路由',
        # 金额相关（精确匹配）
        '破损金额': '金额', '赔偿金额': '金额', '金额(元)': '金额', '金额（元）': '金额',
        '金额元': '金额', '货值': '金额', '价值': '金额', '单价': '金额', '费用总额': '金额',
        '总金额': '金额', '费用': '金额', '赔款': '金额', '赔偿': '金额', '损失': '金额',
        '合计': '金额', '金额合计': '金额', '总计': '金额', '总价': '金额', '总价值': '金额',
        '金额小计': '金额', '实付金额': '金额', '应付金额': '金额', '赔付金额': '金额',
        '赔付款': '金额', '赔偿款': '金额', '扣款金额': '金额',
        # 商品相关
        '商品': '商品详情', '商品名': '商品详情', '商品名称': '商品详情', '货品': '商品详情',
        '品名': '商品详情', '货物': '商品详情',
        # 异常相关
        '异常类型': '异常情况', '异常': '异常情况', '问题类型': '异常情况', '问题': '异常情况',
        # 处理相关
        '处理结果': '处理方式', '处理': '处理方式', '处置': '处理方式',
        # 人员相关
        '负责人': '处理人', '经办人': '处理人', '操作人': '处理人', '经手人': '处理人',
        '处理人员': '处理人',
        # 回款相关
        '回款': '回款情况', '回款状态': '回款情况', '收款': '回款情况', '回款情况说明': '回款情况',
    }
    df = df.rename(columns=column_mapping)

    # ── 智能兜底：日期（优先于金额，避免误判）──
    if '日期' not in df.columns:
        for col in df.columns:
            if any(kw in str(col) for kw in ['日期', '时间']) and \
               not any(kw in str(col) for kw in ['金额', '价', '费', '赔', '损失', '款']):
                df = df.rename(columns={col: '日期'})
                break  # 只映射第一个匹配到的列

    # ── 智能兜底：金额 ──
    if '金额' not in df.columns:
        for col in df.columns:
            if any(kw in str(col) for kw in ['金额', '价', '费', '赔', '损失', '款']):
                df = df.rename(columns={col: '金额'})
                break  # 只映射第一个匹配到的列

    return df

def parse_date_flex(s):
    """尝试多种日期格式解析，失败返回原字符串（模块级，供所有函数共用）"""
    if pd.isna(s):
        return ''
    s = str(s).strip()
    if not s:
        return ''
    # 尝试 pandas 智能解析
    try:
        return pd.to_datetime(s).strftime('%Y-%m-%d')
    except Exception:
        pass
    # 尝试常见格式
    for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%Y年%m月%d日', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return pd.to_datetime(s, format=fmt).strftime('%Y-%m-%d')
        except Exception:
            continue
    # 真的失败了，返回原字符串（后续还能从原始字符串提取年月）
    return s


def _parse_excel_to_records(file_source):
    """将Excel文件源(Base64字符串或文件对象)解析为记录列表

    Args:
        file_source: Base64字符串(可含data:xxx;base64,前缀) 或 Flask FileStorage对象
    Returns:
        list[dict]: 处理后的记录列表
    """
    # 判断输入类型并读取Excel
    if isinstance(file_source, str):
        # Base64字符串
        if ',' in file_source:
            file_source = file_source.split(',')[1]
        file_bytes = base64.b64decode(file_source)
        df = pd.read_excel(io.BytesIO(file_bytes))
    else:
        # Flask FileStorage对象
        df = pd.read_excel(file_source)

    # 标准化列名
    df = standardize_columns(df)

    # ── 日期解析：使用模块级 parse_date_flex ──
    df['日期'] = df['日期'].apply(parse_date_flex)
    df = df.fillna('')
    df = df[df['包裹号'].notna() & (df['包裹号'] != '')]

    # 智能解析金额列（支持 ¥1,000 / 1000元 / 100 等各种格式）
    if '金额' in df.columns:
        df['金额'] = df['金额'].apply(lambda x: parse_amount(x))

    # 其他列转为字符串
    for col in df.columns:
        if col != '金额':
            df[col] = df[col].apply(lambda x: '' if pd.isna(x) or str(x) == 'nan' else str(x))

    records = df.to_dict('records')
    _date_pat = re.compile(r'^\d{4}[-/年]\d{1,2}[-/月]\d{1,2}')
    for _r in records:
        _d = str(_r.get('日期', '') or '').strip()
        _s = str(_r.get('班次', '') or '').strip()
        if not _d and _date_pat.match(_s):
            _r['日期'] = _s
            _r['班次'] = ''
    return records


def _count_filled_fields(item):
    """计算一条记录中非空字段的数量"""
    return sum(1 for v in item.values() if str(v).strip() not in ['', 'nan', 'None'])


def _merge_data(new_data, existing_data):
    """合并新数据与现有数据，重复记录保留更详细的版本
    
    Returns:
        tuple: (final_data, added_count, replaced_count)
    """
    merged = {}
    for item in existing_data:
        key = (str(item.get('日期', '')).strip(), str(item.get('包裹号', '')).strip())
        if key[1]:
            merged[key] = item
    
    added_count = 0
    replaced_count = 0
    for item in new_data:
        key = (str(item.get('日期', '')).strip(), str(item.get('包裹号', '')).strip())
        if key[1]:
            if key in merged:
                # 保留字段更完整的记录
                if _count_filled_fields(item) >= _count_filled_fields(merged[key]):
                    merged[key] = item
                replaced_count += 1
            else:
                merged[key] = item
                added_count += 1
    
    final_data = list(merged.values())
    final_data.sort(key=lambda x: x.get('日期', ''), reverse=True)
    return final_data, added_count, replaced_count


@app.route('/api/import-preview', methods=['POST'])
def import_preview():
    """预览导入数据"""
    try:
        data = request.get_json()
        file_data = data.get('fileData', '')
        if not file_data:
            return jsonify({'success': False, 'error': '没有文件数据'}), 400

        new_data = _parse_excel_to_records(file_data)

        # 计算新增和替换数量（不实际合并）
        existing_data = load_data()
        existing_keys = set()
        for item in existing_data:
            key = (str(item.get('日期', '')).strip(), str(item.get('包裹号', '')).strip())
            existing_keys.add(key)
        
        added = 0
        replaced = 0
        for item in new_data:
            key = (str(item.get('日期', '')).strip(), str(item.get('包裹号', '')).strip())
            if key in existing_keys:
                replaced += 1
            else:
                added += 1
        
        return jsonify({
            'success': True,
            'total': len(new_data),
            'added': added,
            'replaced': replaced,
            'preview': new_data[:5]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/import-confirm', methods=['POST'])
def import_confirm():
    """确认导入数据"""
    try:
        data = request.get_json()
        file_data = data.get('fileData', '')
        if not file_data:
            return jsonify({'success': False, 'error': '没有文件数据'}), 400

        new_data = _parse_excel_to_records(file_data)
        existing_data = load_data()
        final_data, added_count, replaced_count = _merge_data(new_data, existing_data)
        
        save_data(final_data)
        sync_to_github(final_data)
        add_log('导入数据', f'共{len(final_data)}条, 新增{added_count}条, 替换{replaced_count}条')
        
        return jsonify({
            'success': True,
            'total': len(final_data),
            'added': added_count,
            'replaced': replaced_count
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/import', methods=['POST'])
def import_excel():
    """导入Excel文件，支持重复数据替换"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有上传文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '文件名为空'}), 400

        new_data = _parse_excel_to_records(file)
        existing_data = load_data()
        final_data, added_count, replaced_count = _merge_data(new_data, existing_data)
        
        save_data(final_data)
        sync_to_github(final_data)
        add_log('导入数据', f'共{len(final_data)}条, 新增{added_count}条, 替换{replaced_count}条')
        
        return jsonify({
            'success': True,
            'total': len(final_data),
            'added': added_count,
            'replaced': replaced_count
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/batch-paid', methods=['POST'])
def batch_mark_paid():
    """批量标记回款"""
    try:
        indices = json.loads(request.form.get('indices', '[]'))
        data = load_data()
        count = 0
        for idx in indices:
            if 0 <= idx < len(data):
                data[idx]['回款情况'] = '√'
                count += 1
        save_data(data)
        sync_to_github(data)
        return jsonify({'success': True, 'count': count})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/calculate-shared', methods=['POST'])
def calculate_shared_expense():
    """计算公摊金额"""
    import traceback
    print("=== 开始计算公摊 ===")
    try:
        # 处理 FormData 格式
        schedule_file = request.files.get('schedule')
        start_date = request.form.get('start_date', '')
        end_date = request.form.get('end_date', '')
        keywords = request.form.get('keywords', '破损,买赔,赔')
        exclude_responsibility = request.form.get('exclude_resp', '')
        resigned_input = request.form.get('resigned', '')
        resigned_people = [p.strip() for p in resigned_input.split(',') if p.strip()] if resigned_input else []
        
        if not start_date or not end_date:
            return jsonify({'success': False, 'error': '请设置起止日期'}), 400
        
        # 读取排班数据（如果不上传，则从台账的处理人字段提取）
        if schedule_file:
            schedule_df = pd.read_excel(schedule_file)
        else:
            # 从台账的处理人字段提取人员名单
            nc_data_for_schedule = load_data()
            fake_data = []
            for item in nc_data_for_schedule:
                date = str(item.get('日期', '')).strip()
                person = str(item.get('处理人', '')).strip()
                shift = str(item.get('班次', '')).strip()
                if date and person and person not in ['', 'nan']:
                    fake_data.append({
                        '日期': date,
                        '姓名': person,
                        '班次名称': '白班' if '白' in shift else ('夜班' if '夜' in shift else shift)
                    })
            schedule_df = pd.DataFrame(fake_data)
            print(f"未上传排班文件，从台账提取 {len(schedule_df)} 条人员记录")
        schedule_df['日期'] = pd.to_datetime(schedule_df['日期']).dt.strftime('%Y-%m-%d')
        
        # 剔除非全日制合同工
        if '用工性质' in schedule_df.columns:
            before_count = len(schedule_df)
            schedule_df = schedule_df[schedule_df['用工性质'] != '非全日制劳动合同工']
            excluded_workers = before_count - len(schedule_df)
            if excluded_workers > 0:
                print(f'已剔除 {excluded_workers} 条非全日制合同工记录')
        
        # 自动识别列名（兼容不同格式）
        # 姓名列
        name_col = None
        for c in ['姓名', '人员', '处理人', '名字', 'name']:
            if c in schedule_df.columns:
                name_col = c
                break
        # 班次列
        shift_col_name = None
        for c in ['班次名称', '班次', '班型', '排班', 'shift']:
            if c in schedule_df.columns:
                shift_col_name = c
                break
        # 打卡时间列（判断是否实际出勤）
        clock_col_name = None
        for c in ['实际上班时间', '上班打卡', '打卡时间', '签到时间']:
            if c in schedule_df.columns:
                clock_col_name = c
                break
        
        if not name_col:
            return jsonify({'success': False, 'error': f'排班文件缺少姓名列，当前列名: {list(schedule_df.columns)}'}), 400
        if not shift_col_name:
            return jsonify({'success': False, 'error': f'排班文件缺少班次列，当前列名: {list(schedule_df.columns)}'}), 400
        
        # 识别跳班班次列
        jump_shift_col = None
        for c in ['跳班班次', '跳班', 'jump_shift']:
            if c in schedule_df.columns:
                jump_shift_col = c
                break
        
        # 默认跳班班次为中班5次
        default_jump_shift = '中班5次'
        
        # 班次识别函数
        def classify_shift(shift_str, clock_time=None, is_jump_shift=False):
            """识别班次类型: 返回 'day'(白班) 或 'night'(夜班) 或 'rest'(休息) 或 None(未知)
            is_jump_shift: 是否为跳班班次（跳班优先）
            """
            # 先检查是否休息
            if shift_str:
                s = str(shift_str).strip()
                if '休息' in s or '休' == s:
                    return 'rest'
            
            # 检查是否打卡（无打卡时间视为休息/未上班）
            if clock_time is not None:
                ct = str(clock_time).strip()
                if ct in ['', 'nan', 'None', 'null']:
                    return 'rest'  # 没打卡视为休息
            
            if not shift_str or str(shift_str).strip() in ['', 'nan']:
                return None
            s = str(shift_str).strip()
            
            # 跳班班次识别（优先处理）
            if is_jump_shift:
                # 跳班班次固定为夜班
                return 'jump_night'
            
            # 白班: 早班、中班1次、中班3次、中班4次
            if '早班' in s:
                return 'day'
            if '晚班' in s:
                return 'night'
            if '中班' in s:
                import re
                nums = re.findall(r'中班(\d+)次', s)
                if nums:
                    n = int(nums[0])
                    if n in [1, 3, 4]:
                        return 'day'
                    elif n in [2, 5]:
                        return 'night'
            return None  # 无法识别

        # 构建排班字典: {日期: {'day': [白班人员], 'night': [夜班人员], 'jump_night': [跳班人员], 'all': [全部人员]}}
        # 跳班人员单独记录，白班计算时需剔除
        schedule = {}
        
        # 使用打卡文件的实际列名
        name_col = name_col  # 已在前面设置
        shift_col = shift_col_name  # 已在前面设置
        clock_col = clock_col_name  # 已在前面设置
        
        for _, row in schedule_df.iterrows():
            date = str(row.get('日期', '')).strip()
            person = str(row.get(name_col, '')).strip() if name_col else ''
            if not date or not person or person in ['', 'nan', 'None']:
                continue
            if date not in schedule:
                schedule[date] = {'day': [], 'night': [], 'jump_night': [], 'all': []}

            # 获取打卡时间（判断是否实际出勤）
            clock_time = row.get(clock_col) if clock_col else None
            
            # 识别班次（固定班次）
            shift_val = str(row.get(shift_col, '')).strip() if shift_col else ''
            shift_type = classify_shift(shift_val, clock_time, is_jump_shift=False)

            # 识别跳班班次
            jump_shift_val = None
            if jump_shift_col:
                jump_shift_val = str(row.get(jump_shift_col, '')).strip() if jump_shift_col else ''
            if not jump_shift_val or jump_shift_val in ['', 'nan', 'None']:
                jump_shift_val = default_jump_shift  # 默认中班5次
            
            # 判断是否为跳班（跳班班次名称匹配）
            is_jump = shift_val == jump_shift_val
            jump_shift_type = classify_shift(shift_val, clock_time, is_jump_shift=is_jump)

            # 跳过休息的人
            if shift_type == 'rest' and jump_shift_type == 'rest':
                continue

            # 记录到all列表
            if person not in schedule[date]['all']:
                schedule[date]['all'].append(person)

            # 处理跳班班次（优先级最高）
            if jump_shift_type == 'jump_night':
                # 跳班为夜班，加入jump_night列表
                if person not in schedule[date]['jump_night']:
                    schedule[date]['jump_night'].append(person)
                # 如果已在白班列表，移除（跳班优先）
                if person in schedule[date]['day']:
                    schedule[date]['day'].remove(person)
                # 同时加入夜班列表参与公摊
                if person not in schedule[date]['night']:
                    schedule[date]['night'].append(person)
            elif shift_type == 'day':
                # 固定班次为白班
                # 如果此人当天有跳班，则不加入白班
                if person not in schedule[date]['jump_night']:
                    if person not in schedule[date]['day']:
                        schedule[date]['day'].append(person)
            elif shift_type == 'night':
                # 固定班次为夜班
                if person not in schedule[date]['night']:
                    schedule[date]['night'].append(person)
            else:
                # 无法识别班次，归入全部
                pass
        
        # 剔除离职人员（从排班中移除）
        if resigned_people:
            print(f"需剔除离职人员: {resigned_people}")
            removed_count = 0
            for date in list(schedule.keys()):
                info = schedule[date]
                for fmt in info:
                    if isinstance(info, dict):
                        if isinstance(info[fmt], list):
                            orig_len = len(info[fmt])
                            info[fmt] = [p for p in info[fmt] if p not in resigned_people]
                            removed_count += orig_len - len(info[fmt])
            print(f"已从排班中剔除 {removed_count} 条离职人员记录")
        
        # 自动识别单责任人（具体人名）- 包含以下特征的视为单责：
        # 1. 责任方只包含一个具体人名（如：张景莉、吴光辉）
        # 2. 不包含"共责"、"NC"、"验货"、"卸车"等共同责任关键词
        # 从排班文件中提取所有真实人名
        all_persons_in_schedule = set()
        for date_info in schedule.values():
            for p in date_info.get('day', []):
                all_persons_in_schedule.add(p)
            for p in date_info.get('night', []):
                all_persons_in_schedule.add(p)
        
        def classify_responsibility(resp):
            """
            分析责任方，返回 (action, ratio)
            action:
              'exclude'  — 整条剔除（单人全责 或 两个真实人名共责）
              'half'     — 金额减半后参与公摊（一个真实人名 + NC/共责等）
              'include'  — 正常参与公摊
            ratio: 参与公摊的金额比例（1.0 或 0.5）
            """
            if not resp or resp == '':
                return ('include', 1.0)

            # 统计责任方中包含多少个排班真实人名
            matched_persons = [p for p in all_persons_in_schedule if p in resp]
            n = len(matched_persons)

            if n == 0:
                # 没有真实人名：未拦截、NC、卸车/NC共责 等 → 正常公摊
                return ('include', 1.0)
            elif n == 1:
                # 恰好一个真实人名
                person = matched_persons[0]
                if resp == person:
                    # 完全匹配：单人全责 → 剔除
                    return ('exclude', 0.0)
                else:
                    # 含有其他内容（如"/NC共责"、"&NC"）→ 半责，金额÷2参与公摊
                    return ('half', 0.5)
            else:
                # 两个及以上真实人名（如"张三&李四"）→ 剔除，但人员仍参与公摊
                return ('exclude', 0.0)
        
        # 筛选时间范围内的破损买赔数据
        nc_data = load_data()
        damaged_items = []
        excluded_items = []  # 被剔除的单责记录
        keyword_list = [k.strip() for k in keywords.split(',') if k.strip()]
        
        print(f"台账数据: {len(nc_data)} 条")
        print(f"关键词: {keyword_list}")
        
        for item in nc_data:
            item_date = str(item.get('日期', '')).strip()
            exception_type = str(item.get('异常情况', ''))
            responsibility = str(item.get('责任方', '')).strip()
            
            # 识别破损买赔相关单子
            is_damaged = any(keyword in exception_type for keyword in keyword_list) if keyword_list else True
            is_in_range = start_date <= item_date <= end_date
            has_amount = item.get('金额') is not None and item.get('金额') != ''
            
            if is_damaged and is_in_range and has_amount:
                amount = parse_amount(item.get('金额'))
                if amount > 0:
                    action, ratio = classify_responsibility(responsibility)
                    if action == 'exclude':
                        excluded_items.append({
                            'date': item_date,
                            'package': item.get('包裹号', ''),
                            'responsibility': responsibility,
                            'amount': amount
                        })
                    elif action == 'half':
                        # 金额减半后参与公摊
                        half_item = dict(item)
                        half_item['金额'] = round(amount * 0.5, 2)
                        half_item['_original_amount'] = amount
                        half_item['_half_note'] = f'{responsibility}（半责，原{amount}元÷2）'
                        damaged_items.append(half_item)
                    else:
                        damaged_items.append(item)
        
        # 逐单计算公摊
        results = {}
        daily_details = {}
        
        print(f"破损条目: {len(damaged_items)} 条")
        print(f"排除条目: {len(excluded_items)} 条")
        
        # 逐单计算公摊：每一单破损按其班次的当班人数均摊
        item_no = 0
        for item in damaged_items:
            item_no += 1
            date = str(item.get('日期', '')).strip()
            shift = str(item.get('班次', '')).strip()
            amount = parse_amount(item.get('金额'))
            if amount <= 0:
                continue
            day_info = schedule.get(date, {})
            if isinstance(day_info, list):
                day_people, night_people = [], []
            else:
                day_people = day_info.get('day', [])
                night_people = day_info.get('night', [])
            
            # 判断该单属于白班还是夜班
            if '夜' in shift:
                people = night_people
                shift_label = '夜班'
            else:
                people = day_people
                shift_label = '白班'
            
            if not people:
                continue
            
            per_person = amount / len(people)
            detail_key = f"{date}_{shift_label}_单{item_no}"
            daily_details[detail_key] = {
                'total': round(amount, 2),
                'people': len(people),
                'per_person': round(per_person, 2),
                'person_list': people,
                'shift_label': shift_label,
                'package': item.get('包裹号', ''),
                'responsibility': item.get('_half_note', str(item.get('责任方', '')))
            }
            for person in people:
                if person not in results:
                    results[person] = {'total': 0, 'dates': []}
                results[person]['total'] = round(results[person]['total'] + per_person, 2)
                results[person]['dates'].append({
                    'date': date,
                    'shift': shift_label,
                    'package': item.get('包裹号', ''),
                    'amount': round(per_person, 2)
                })
        
        # 汇总排序
        summary = [{'person': p, 'total': d['total'], 'details': d['dates']} 
                   for p, d in results.items()]
        summary.sort(key=lambda x: x['total'], reverse=True)
        
        grand_total = sum(r['total'] for r in summary)
        
        return jsonify({
            'success': True,
            'start_date': start_date,
            'end_date': end_date,
            'total_damaged': round(sum(parse_amount(i.get('金额')) for i in damaged_items), 2),
            'days_count': len(set(str(i.get('日期','')) for i in damaged_items)),
            'people_count': len(results),
            'grand_total': round(grand_total, 2),
            'summary': summary,
            'daily_details': json.loads(json.dumps({k: {kk: (str(vv) if isinstance(vv, (np.integer, np.floating)) else list(vv) if isinstance(vv, (set, frozenset)) else (float(vv) if isinstance(vv, (int, float)) and not isinstance(vv, bool) else vv)) for kk, vv in v.items()} for k, v in daily_details.items()}, default=str)),
            'excluded_count': len(excluded_items),
            'excluded_list': excluded_items,
            'half_list': [{'date': str(i.get('日期','')), 'package': i.get('包裹号',''), 'responsibility': i.get('责任方',''), 'original_amount': i.get('_original_amount', i.get('金额',0)), 'half_amount': i.get('金额',0)} for i in damaged_items if i.get('_half_note')],
            'resigned_people': resigned_people
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export-shared', methods=['POST'])
def export_shared_expense():
    """导出公摊计算结果为Excel"""
    try:
        # 接收计算结果数据
        data = request.json or {}
        summary = data.get('summary', [])
        daily_details = data.get('daily_details', {})
        excluded_list = data.get('excluded_list', [])
        start_date = data.get('start_date', '')
        end_date = data.get('end_date', '')
        
        # 预先按人名汇总白班/夜班金额
        person_day_total = {}   # {人名: 白班合计}
        person_night_total = {} # {人名: 夜班合计}
        for key, info in daily_details.items():
            per_person = info.get('per_person', 0)
            if '白班' in key:
                for p in info.get('day_persons', []):
                    person_day_total[p] = round(person_day_total.get(p, 0) + per_person, 2)
            elif '夜班' in key:
                for p in info.get('night_persons', []):
                    person_night_total[p] = round(person_night_total.get(p, 0) + per_person, 2)

        # 创建结果DataFrame（含白班/夜班小计）
        rows = []
        for item in summary:
            day_amt = person_day_total.get(item['person'], 0)
            night_amt = person_night_total.get(item['person'], 0)
            day_details_str  = ', '.join([d['date']+'('+str(d['amount'])+')' for d in item['details'] if d.get('shift')=='白班'])
            night_details_str = ', '.join([d['date']+'('+str(d['amount'])+')' for d in item['details'] if d.get('shift')=='夜班'])
            rows.append({
                '姓名': item['person'],
                '总公摊金额': item['total'],
                '白班合计': day_amt if day_amt else '',
                '夜班合计': night_amt if night_amt else '',
                '涉及天数': len(item['details']),
                '白班明细': day_details_str,
                '夜班明细': night_details_str,
            })

        result_df = pd.DataFrame(rows)
        
        # 创建白班明细和夜班明细（按人名分组）
        # 先按班次分开处理
        day_details = {k: v for k, v in daily_details.items() if '白班' in k}
        night_details = {k: v for k, v in daily_details.items() if '夜班' in k}
        
        # 白班明细：按人名分组
        day_person_details = {}  # {人名: {日期: 金额}}
        all_day_dates = sorted(set(k.split('_')[0] for k in day_details.keys()))
        for key, info in day_details.items():
            date = key.split('_')[0]
            people = info.get('person_list', [])
            per_person = info['per_person']
            
            for person in people:
                if person not in day_person_details:
                    day_person_details[person] = {'总公摊': 0, '天数': 0}
                day_person_details[person][date] = day_person_details[person].get(date, 0) + per_person
                day_person_details[person]['总公摊'] += per_person
                day_person_details[person]['天数'] += 1
        
        # 构建白班DataFrame
        day_rows = []
        for person, details in sorted(day_person_details.items(), key=lambda x: -x[1]['总公摊']):
            row = {'姓名': person, '总公摊': round(details['总公摊'], 2), '天数': details['天数']}
            for date in all_day_dates:
                if date in details:
                    row[date] = details[date]
            day_rows.append(row)
        day_df = pd.DataFrame(day_rows)
        
        # 夜班明细：按人名分组
        night_person_details = {}
        all_night_dates = sorted(set(k.split('_')[0] for k in night_details.keys()))
        for key, info in night_details.items():
            date = key.split('_')[0]
            people = info.get('person_list', [])
            per_person = info['per_person']
            
            for person in people:
                if person not in night_person_details:
                    night_person_details[person] = {'总公摊': 0, '天数': 0}
                night_person_details[person][date] = night_person_details[person].get(date, 0) + per_person
                night_person_details[person]['总公摊'] += per_person
                night_person_details[person]['天数'] += 1
        
        # 构建夜班DataFrame
        night_rows = []
        for person, details in sorted(night_person_details.items(), key=lambda x: -x[1]['总公摊']):
            row = {'姓名': person, '总公摊': round(details['总公摊'], 2), '天数': details['天数']}
            for date in all_night_dates:
                if date in details:
                    row[date] = details[date]
            night_rows.append(row)
        night_df = pd.DataFrame(night_rows)
        
        # 被剔除的记录
        excluded_df = pd.DataFrame(excluded_list) if excluded_list else pd.DataFrame()
        
        # 写入Excel（多个Sheet）
        excel_path = '/tmp/nc_shared_expense.xlsx'
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            if not result_df.empty:
                result_df.to_excel(writer, index=False, sheet_name='公摊汇总')
            if not day_df.empty:
                day_df.to_excel(writer, index=False, sheet_name='白班明细')
            if not night_df.empty:
                night_df.to_excel(writer, index=False, sheet_name='夜班明细')
            if not excluded_df.empty:
                excluded_df.to_excel(writer, index=False, sheet_name='已剔除记录')
        
        return send_file(excel_path, as_attachment=True, download_name='公摊计算结果.xlsx')
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Render.com 需要


@app.route('/api/export-template', methods=['POST'])
def export_template():
    """按模板导出数据"""
    try:
        req = request.get_json()
        export_type = req.get('type', 'all')
        data = req.get('data', [])
        
        if not data:
            return jsonify({'success': False, 'error': '没有数据'}), 400
        
        df = pd.DataFrame(data)
        
        # 根据导出类型设置列名
        if export_type == 'by-resp':
            # 责任方汇总
            columns = ['责任方', '数量', '金额']
            df = df[columns] if all(c in df.columns for c in columns) else df
        else:
            # 其他类型保持原列
            preferred_columns = ['日期', '班次', '包裹号', '商品详情', '异常情况', '金额', '责任方', '处理方式', '路由', '处理人', '回款情况']
            existing_cols = [c for c in preferred_columns if c in df.columns]
            df = df[existing_cols] if existing_cols else df
        
        # 创建Excel
        excel_path = '/tmp/nc_export.xlsx'
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='数据')
            
            # 获取工作表并设置列宽
            ws = writer.sheets['数据']
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
        
        return send_file(excel_path, as_attachment=True, download_name=f'{export_type}_{datetime.now(timezone.utc).strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500




# ==================== 定时同步功能 ====================

# 存储上次同步时间
LAST_SYNC_FILE = '/tmp/last_sync_time.json'

def get_last_sync_time():
    """获取上次同步时间"""
    try:
        if os.path.exists(LAST_SYNC_FILE):
            with open(LAST_SYNC_FILE, 'r') as f:
                return json.load(f).get('last_sync', '')
    except:
        pass
    return ''

def save_last_sync_time():
    """保存同步时间"""
    try:
        with open(LAST_SYNC_FILE, 'w') as f:
            json.dump({'last_sync': datetime.now(timezone.utc).isoformat()}, f)
    except:
        pass

@app.route('/api/sync-status')
def sync_status():
    """获取同步状态"""
    return jsonify({
        'last_sync': get_last_sync_time(),
        'data_count': len(load_data()),
        'server_time': datetime.now(timezone.utc).isoformat()
    })

@app.route('/api/manual-sync', methods=['POST'])
def manual_sync():
    """手动触发同步"""
    try:
        data = load_data()
        success, err_msg = sync_to_github(data)
        if success:
            save_last_sync_time()
            return jsonify({'success': True, 'message': '同步成功', 'data_count': len(data)})
        else:
            return jsonify({'success': False, 'error': err_msg}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/schedule-sync', methods=['POST'])
def schedule_sync():
    """设置定时同步（需要外部cron或Render的cron job调用）"""
    try:
        data = load_data()
        sync_to_github(data)
        save_last_sync_time()
        return jsonify({'success': True, 'synced_at': datetime.now(timezone.utc).isoformat()})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500




# ==================== 自动备份功能 ====================

BACKUP_DIR = '/tmp/backups'
MAX_BACKUPS = 30  # 保留最近30天备份

def ensure_backup_dir():
    """确保备份目录存在"""
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)

def create_backup():
    """创建数据备份"""
    try:
        ensure_backup_dir()
        
        data = load_data()
        timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
        backup_file = os.path.join(BACKUP_DIR, f'backup_{timestamp}.json')
        
        with open(backup_file, 'w', encoding='utf-8') as f:
            json.dump({
                'timestamp': datetime.now(timezone.utc).isoformat(),
                'data_count': len(data),
                'data': data
            }, f, ensure_ascii=False, indent=2)
        
        # 清理旧备份（保留最近30天）
        clean_old_backups()
        
        add_log('自动备份', f'备份成功: {len(data)}条数据')
        return True
    except Exception as e:
        print(f'备份失败: {e}')
        return False

def clean_old_backups():
    """清理旧备份"""
    try:
        ensure_backup_dir()
        files = sorted([f for f in os.listdir(BACKUP_DIR) if f.startswith('backup_')])
        
        while len(files) > MAX_BACKUPS:
            old_file = os.path.join(BACKUP_DIR, files[0])
            os.remove(old_file)
            files.pop(0)
    except Exception as e:
        print(f'清理旧备份失败: {e}')

def get_backup_list():
    """获取备份列表"""
    try:
        ensure_backup_dir()
        files = sorted([f for f in os.listdir(BACKUP_DIR) if f.startswith('backup_')], reverse=True)
        
        backups = []
        for f in files:
            filepath = os.path.join(BACKUP_DIR, f)
            stat = os.stat(filepath)
            backups.append({
                'filename': f,
                'size': stat.st_size,
                'time': datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()
            })
        return backups
    except:
        return []

@app.route('/api/backups')
def api_backups():
    """获取备份列表"""
    backups = get_backup_list()
    return jsonify({'success': True, 'backups': backups})

@app.route('/api/backup-now', methods=['POST'])
def api_backup_now():
    """立即创建备份"""
    success = create_backup()
    if success:
        return jsonify({'success': True, 'message': '备份成功'})
    return jsonify({'success': False, 'error': '备份失败'}), 500

@app.route('/api/restore-backup/<filename>', methods=['POST'])
def api_restore_backup(filename):
    """恢复备份"""
    try:
        filepath = os.path.join(BACKUP_DIR, filename)
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'error': '备份文件不存在'}), 404
        
        with open(filepath, 'r') as f:
            backup_data = json.load(f)
        
        data = backup_data.get('data', [])
        save_data(data)
        sync_to_github(data)
        add_log('恢复备份', f'从 {filename} 恢复 {len(data)} 条数据')
        
        return jsonify({'success': True, 'message': f'已恢复 {len(data)} 条数据'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/download-backup/<filename>')
def api_download_backup(filename):
    """下载备份文件"""
    try:
        filepath = os.path.join(BACKUP_DIR, filename)
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500




# ==================== 数据分析报告功能 ====================

ANALYSIS_DIR = '/tmp/analysis_reports'
os.makedirs(ANALYSIS_DIR, exist_ok=True)

@app.route('/analysis')
def analysis_page():
    """数据分析报告页面"""
    return render_template('analysis.html')

@app.route('/api/analysis/upload', methods=['POST'])
def api_analysis_upload():
    """上传历史数据（不存入主数据，只用于分析）"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请选择文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '请选择文件'}), 400
        
        # 保存临时文件
        filename = f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        filepath = os.path.join(ANALYSIS_DIR, filename)
        file.save(filepath)
        
        # 读取数据，提取年份
        df = pd.read_excel(filepath)
        df.columns = df.columns.str.strip()
        
        # 尝试从"日期"列提取年份
        years = []
        if '日期' in df.columns:
            try:
                df['日期'] = pd.to_datetime(df['日期'], errors='coerce')
                years = df['日期'].dt.year.dropna().unique().tolist()
            except:
                pass
        
        return jsonify({
            'success': True,
            'filename': filename,
            'years': years,
            'record_count': len(df)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/analysis/generate', methods=['POST'])
def api_analysis_generate():
    """生成数据分析报告"""
    try:
        data = request.get_json()
        filenames = data.get('filenames', [])  # 上传的历史数据文件
        include_current = data.get('include_current', True)  # 是否包含当前系统数据
        
        print(f"[DEBUG] 收到生成报告请求: filenames={filenames}, include_current={include_current}")
        
        # 收集所有数据
        all_data = []
        
        # 1. 读取上传的历史数据（复用台账系统完整解析逻辑）
        for filename in filenames:
            filepath = os.path.join(ANALYSIS_DIR, filename)
            if os.path.exists(filepath):
                print(f"[DEBUG] 读取历史数据文件: {filepath}")
                df = pd.read_excel(filepath)
                # ── 关键修复：复用台账系统的标准化逻辑 ──
                df = standardize_columns(df)  # 列名标准化（时间→日期、单号→凭证等）
                if '日期' in df.columns:
                    df['日期'] = df['日期'].apply(parse_date_flex)  # 日期解析
                if '金额' in df.columns:
                    df['金额'] = df['金额'].apply(parse_amount)  # 金额解析
                all_data.append(df)
                print(f"[DEBUG] 历史数据条数: {len(df)}")

        # 2. 读取当前系统数据（带金额智能解析）
        if include_current:
            print(f"[DEBUG] 读取当前系统数据...")
            current_data = load_data()
            if current_data:
                df_current = pd.DataFrame(current_data)
                if '金额' in df_current.columns:
                    df_current['金额'] = df_current['金额'].apply(parse_amount)
                all_data.append(df_current)
                print(f"[DEBUG] 当前系统数据条数: {len(df_current)}")
        
        if not all_data:
            return jsonify({'success': False, 'error': '没有数据可用于分析'}), 400
        
        # 合并所有数据
        df_all = pd.concat(all_data, ignore_index=True)
        print(f"[DEBUG] 合并后总数据条数: {len(df_all)}")
        print(f"[DEBUG] 数据列名: {list(df_all.columns)}")
        
        # 生成报告（使用 openpyxl）
        report_filename = f"数据分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = os.path.join(ANALYSIS_DIR, report_filename)
        print(f"[DEBUG] 报告将保存到: {report_path}")
        
        # 调用报告生成函数
        generate_analysis_report(df_all, report_path)
        print(f"[DEBUG] 报告生成成功: {report_filename}")
        
        return jsonify({
            'success': True,
            'report_filename': report_filename,
            'download_url': f'/api/analysis/download/{report_filename}'
        })
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"[ERROR] 生成报告失败: {str(e)}")
        print(f"[ERROR] 详细堆栈: {error_trace}")
        return jsonify({'success': False, 'error': str(e), 'traceback': error_trace}), 500

@app.route('/api/analysis/download/<filename>')
def api_analysis_download(filename):
    """下载分析报告"""
    try:
        filepath = os.path.join(ANALYSIS_DIR, filename)
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'error': '文件不存在'}), 404
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/analysis/generate-pdf', methods=['POST'])
def api_analysis_generate_pdf():
    """生成精美的 PDF 分析报告"""
    try:
        data = request.get_json()
        filenames = data.get('filenames', [])
        include_current = data.get('include_current', True)

        print(f"[PDF] 收到PDF生成请求: filenames={filenames}, include_current={include_current}")

        # 收集数据（复用台账系统完整解析逻辑）
        all_data = []
        for filename in filenames:
            filepath = os.path.join(ANALYSIS_DIR, filename)
            if os.path.exists(filepath):
                df = pd.read_excel(filepath)
                df.columns = df.columns.str.strip()
                df = standardize_columns(df)  # 列名标准化（时间→日期、单号→凭证等）
                if '日期' in df.columns:
                    df['日期'] = df['日期'].apply(parse_date_flex)  # 日期解析
                if '金额' in df.columns:
                    df['金额'] = df['金额'].apply(parse_amount)  # 金额解析
                all_data.append(df)

        if include_current:
            current_data = load_data()
            if current_data:
                df_cur = pd.DataFrame(current_data)
                if '金额' in df_cur.columns:
                    df_cur['金额'] = df_cur['金额'].apply(parse_amount)
                all_data.append(df_cur)

        if not all_data:
            return jsonify({'success': False, 'error': '没有数据可用于分析'}), 400

        df_all = pd.concat(all_data, ignore_index=True)
        print(f"[PDF] 合并后总数据: {len(df_all)} 条")

        # 生成精美 HTML 报告（浏览器 Ctrl+P 保存为 PDF）
        report_filename = f"分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"

        html_content = render_analysis_html_report(df_all)
        print(f"[PDF] HTML报告生成成功: {report_filename} ({len(html_content)} 字节)")

        return jsonify({
            'success': True,
            'html': html_content,
            'filename': report_filename
        })

    except Exception as e:
        import traceback
        print(f"[PDF ERROR] {str(e)}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500

def generate_analysis_report(df, output_path):
    """生成数据分析报告（Excel格式），含年同比对比"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter

    # ========== 数据预处理 ==========
    # ── 日期解析：统一使用 parse_date_flex（与台账系统一致）──
    if '日期' in df.columns:
        df['日期'] = df['日期'].apply(lambda x: parse_date_flex(x) if pd.notna(x) else '')
        # 将解析后的日期字符串转为 datetime
        df['日期'] = pd.to_datetime(df['日期'], errors='coerce')
        df['年份'] = df['日期'].dt.year.fillna(0).astype(int)
        df['月份'] = df['日期'].dt.month.fillna(0).astype(int)
        df['年月'] = df['日期'].dt.to_period('M').astype(str)
        df['年月'] = df['年月'].replace('NaT', '未知')
    else:
        df['年份'] = 0
        df['月份'] = 0
        df['年月'] = '未知'
    
    if '金额' in df.columns:
        df['金额'] = pd.to_numeric(df['金额'], errors='coerce').fillna(0)

    # 样式定义
    hdr_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    pct_fmt = '0.0%'

    wb = openpyxl.Workbook()

    # ========================================================================
    # Sheet 1: 总览
    # ========================================================================
    ws = wb.active
    ws.title = "总览"
    ws['A1'] = "NC台账数据分析报告"
    ws['A1'].font = title_font
    ws['A3'] = f"生成时间: {datetime.now(timezone(timedelta(hours=8))).strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'] = f"数据总条数: {len(df)}"

    # 获取年份列表
    year_list = sorted(df['年份'].unique())
    valid_years = [y for y in year_list if y > 2000]

    row = 6
    ws[f'A{row}'] = "核心指标"
    ws[f'A{row}'].font = hdr_font
    row += 1

    # 时间范围
    valid_dates = df['日期'].dropna()
    if len(valid_dates) > 0:
        ws[f'A{row}'] = "数据时间范围"
        ws[f'B{row}'] = f"{valid_dates.min().strftime('%Y-%m-%d')} 至 {valid_dates.max().strftime('%Y-%m-%d')}"
        ws[f'C{row}'] = f"覆盖 {len(valid_years)} 个年份"
        row += 1

    # 年份对比表头
    if len(valid_years) >= 2:
        row += 1
        ws[f'A{row}'] = "指标"
        for idx, y in enumerate(valid_years):
            ws.cell(row=row, column=2+idx).value = f"{int(y)}年"
            ws.cell(row=row, column=2+idx).font = hdr_font
        last_col = 2 + len(valid_years)
        ws.cell(row=row, column=last_col).value = "同比变化"
        ws.cell(row=row, column=last_col).font = hdr_font
        row += 1

        # 各年份数据
        year_data = {}
        for y in valid_years:
            d = df[df['年份'] == y]
            year_data[y] = {
                '条数': len(d),
                '总金额': d['金额'].sum(),
                '有金额': (d['金额'] > 0).sum(),
                '平均金额': d['金额'].mean(),
                '责任方数': len(d['责任方'].unique()) if '责任方' in d.columns else 0,
            }

        metrics = [
            ('数据条数', '条数', '{:.0f}'),
            ('破损总金额', '总金额', '¥{:.2f}'),
            ('有金额记录', '有金额', '{:.0f}'),
            ('单均金额', '平均金额', '¥{:.2f}'),
            ('涉及责任方', '责任方数', '{:.0f}'),
        ]

        for label, key, fmt in metrics:
            ws.cell(row=row, column=1).value = label
            vals = []
            for idx, y in enumerate(valid_years):
                val = year_data[y][key]
                ws.cell(row=row, column=2+idx).value = val
                vals.append(val)
            # 同比变化（最近两个年份）
            if len(vals) >= 2 and vals[0] != 0:
                change = (vals[1] - vals[0]) / vals[0]
                ws.cell(row=row, column=last_col).value = change
                ws.cell(row=row, column=last_col).number_format = '0.0%'
            row += 1

        # 同比文字总结
        row += 1
        y_prev, y_curr = valid_years[-2], valid_years[-1]
        d_prev, d_curr = year_data[y_prev], year_data[y_curr]
        cnt_change = (d_curr['条数'] - d_prev['条数']) / d_prev['条数'] * 100
        amt_change = (d_curr['总金额'] - d_prev['总金额']) / d_prev['总金额'] * 100
        ws[f'A{row}'] = f"▶ {int(y_curr)}年 vs {int(y_prev)}年: 条数 {'↑' if cnt_change>0 else '↓'}{abs(cnt_change):.1f}%, 金额 {'↑' if amt_change>0 else '↓'}{abs(amt_change):.1f}%"
        ws[f'A{row}'].font = Font(bold=True, color='2563EB')

    else:
        # 只有一年，简单统计
        row += 1
        ws[f'A{row}'] = "数据条数"
        ws[f'B{row}'] = len(df)
        row += 1
        ws[f'A{row}'] = "破损总金额"
        ws[f'B{row}'] = f"¥{df['金额'].sum():.2f}"
        row += 1
        ws[f'A{row}'] = "有金额记录"
        ws[f'B{row}'] = int((df['金额'] > 0).sum())

    # ========================================================================
    # Sheet 2: 年份同比对比（核心）
    # ========================================================================
    ws_comp = wb.create_sheet("年份对比")
    ws_comp['A1'] = "年份同比对比分析"
    ws_comp['A1'].font = title_font

    if len(valid_years) >= 2:
        y_prev, y_curr = valid_years[-2], valid_years[-1]

        # --- 月度对比 ---
        row = 3
        ws_comp[f'A{row}'] = "月度对比"
        ws_comp[f'A{row}'].font = hdr_font
        row += 1

        ws_comp[f'A{row}'] = "月份"
        ws_comp[f'B{row}'] = f"{int(y_prev)}年"
        ws_comp[f'C{row}'] = f"{int(y_curr)}年"
        ws_comp[f'D{row}'] = "同比变化"
        ws_comp[f'D{row}'].font = hdr_font
        row += 1

        monthly_comp = []
        for m in range(1, 13):
            d_prev_m = df[(df['年份'] == y_prev) & (df['月份'] == m)]
            d_curr_m = df[(df['年份'] == y_curr) & (df['月份'] == m)]
            amt_prev = d_prev_m['金额'].sum()
            amt_curr = d_curr_m['金额'].sum()
            if amt_prev > 0 or amt_curr > 0:
                change = (amt_curr - amt_prev) / amt_prev if amt_prev > 0 else None
                monthly_comp.append((m, amt_prev, amt_curr, change))
                ws_comp[f'A{row}'] = f"{m}月"
                ws_comp[f'B{row}'] = float(amt_prev)
                ws_comp[f'C{row}'] = float(amt_curr)
                if change is not None:
                    ws_comp[f'D{row}'] = change
                    ws_comp[f'D{row}'].number_format = '0.0%'
                row += 1

        # 月度对比图表
        if len(monthly_comp) >= 2:
            chart = BarChart()
            chart.type = "col"  # 纵向柱状图
            chart.title = f"月度破损金额对比 ({int(y_prev)}年 vs {int(y_curr)}年)"
            chart.style = 10
            chart.y_axis.title = "金额 (¥)"
            chart.x_axis.title = "月份"

            data_start = 4  # row of first data
            data_end = data_start + len(monthly_comp) - 1
            ref_prev = Reference(ws_comp, min_col=2, min_row=3, max_row=data_end)
            ref_curr = Reference(ws_comp, min_col=3, min_row=3, max_row=data_end)
            ref_cats = Reference(ws_comp, min_col=1, min_row=data_start, max_row=data_end)

            chart.add_data(ref_prev, titles_from_data=True)
            chart.add_data(ref_curr, titles_from_data=True)
            chart.set_categories(ref_cats)
            # 设置颜色
            from openpyxl.chart.series import DataPoint
            from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
            chart.series[0].graphicalProperties.line.solidFill = "2563EB"  # 蓝色-2024
            chart.series[1].graphicalProperties.line.solidFill = "DC2626"  # 红色-2025
            chart.width = 20
            chart.height = 12
            ws_comp.add_chart(chart, "F3")

        # --- 责任方对比 ---
        if '责任方' in df.columns:
            row += 2
            ws_comp[f'A{row}'] = "责任方对比（按金额）"
            ws_comp[f'A{row}'].font = hdr_font
            row += 1

            ws_comp[f'A{row}'] = "责任方"
            ws_comp[f'B{row}'] = f"{int(y_prev)}年"
            ws_comp[f'C{row}'] = f"{int(y_curr)}年"
            ws_comp[f'D{row}'] = "变化"
            row += 1

            df_prev = df[df['年份'] == y_prev]
            df_curr = df[df['年份'] == y_curr]
            all_parties = set(df_prev['责任方'].unique()) | set(df_curr['责任方'].unique())
            # 过滤掉 NaN/float，避免 sorted() 时 float 与 str 无法比较
            all_parties = {p for p in all_parties if p == p and p is not None}

            party_comp = []
            for p in sorted(all_parties):
                amt_p = df_prev[df_prev['责任方'] == p]['金额'].sum()
                amt_c = df_curr[df_curr['责任方'] == p]['金额'].sum()
                if amt_p > 0 or amt_c > 0:
                    change = (amt_c - amt_p) / amt_p if amt_p > 0 else None
                    party_comp.append((p, amt_p, amt_c, change))
                    ws_comp[f'A{row}'] = str(p)
                    ws_comp[f'B{row}'] = float(amt_p) if amt_p > 0 else 0
                    ws_comp[f'C{row}'] = float(amt_c) if amt_c > 0 else 0
                    if change is not None:
                        ws_comp[f'D{row}'] = change
                        ws_comp[f'D{row}'].number_format = '0.0%'
                    row += 1

    # ========================================================================
    # Sheet 3: 月度趋势
    # ========================================================================
    ws_trend = wb.create_sheet("月度趋势")
    if '日期' in df.columns and '金额' in df.columns:
        monthly_amount = df.groupby('年月')['金额'].sum().reset_index()
        monthly_amount.columns = ['年月', '破损金额']

        ws_trend['A1'] = "月度破损金额统计"
        ws_trend['A1'].font = hdr_font

        ws_trend['A3'] = "年月"
        ws_trend['B3'] = "破损金额"
        for i, (_, r) in enumerate(monthly_amount.iterrows(), start=4):
            ws_trend[f'A{i}'] = str(r['年月'])
            ws_trend[f'B{i}'] = float(r['破损金额'])

        if len(monthly_amount) >= 2:
            chart = LineChart()
            chart.title = "月度破损金额趋势"
            chart.style = 10
            chart.y_axis.title = "金额 (¥)"
            chart.x_axis.title = "年月"
            data_ref = Reference(ws_trend, min_col=2, min_row=3, max_row=3+len(monthly_amount))
            cat_ref = Reference(ws_trend, min_col=1, min_row=4, max_row=3+len(monthly_amount))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cat_ref)
            chart.width = 22
            chart.height = 13
            ws_trend.add_chart(chart, "D3")

    # ========================================================================
    # Sheet 4: 责任方分析
    # ========================================================================
    ws_resp = wb.create_sheet("责任方分析")
    if '责任方' in df.columns and '金额' in df.columns:
        resp_stats = df.groupby('责任方').agg(
            破损次数=('日期', 'count'),
            破损金额=('金额', 'sum'),
            单均金额=('金额', 'mean')
        ).reset_index().sort_values('破损金额', ascending=False)

        ws_resp['A1'] = "责任方综合统计"
        ws_resp['A1'].font = hdr_font
        ws_resp['A3'] = "责任方"
        ws_resp['B3'] = "破损次数"
        ws_resp['C3'] = "破损金额"
        ws_resp['D3'] = "单均金额"

        for i, (_, r) in enumerate(resp_stats.iterrows(), start=4):
            ws_resp[f'A{i}'] = str(r['责任方'])
            ws_resp[f'B{i}'] = int(r['破损次数'])
            ws_resp[f'C{i}'] = float(r['破损金额'])
            ws_resp[f'D{i}'] = float(r['单均金额'])

        if len(resp_stats) >= 2:
            chart = BarChart()
            chart.title = "责任方破损金额 TOP"
            chart.style = 10
            chart.type = "bar"  # 横向条形图
            # 显示前15
            n = min(15, len(resp_stats))
            data_ref = Reference(ws_resp, min_col=3, min_row=3, max_row=3+n)
            cat_ref = Reference(ws_resp, min_col=1, min_row=4, max_row=3+n)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cat_ref)
            chart.width = 22
            chart.height = 14
            ws_resp.add_chart(chart, "F3")

    # ========================================================================
    # Sheet 5: 产品破损排行
    # ========================================================================
    ws_prod = wb.create_sheet("产品破损排行")
    if '商品详情' in df.columns and '金额' in df.columns:
        prod_stats = df.groupby('商品详情').agg(
            破损次数=('金额', 'count'),
            破损金额=('金额', 'sum')
        ).reset_index().sort_values('破损金额', ascending=False).head(20)

        ws_prod['A1'] = "产品破损排行 TOP20"
        ws_prod['A1'].font = hdr_font
        ws_prod['A3'] = "排名"
        ws_prod['B3'] = "商品详情"
        ws_prod['C3'] = "破损次数"
        ws_prod['D3'] = "破损金额"

        for i, (_, r) in enumerate(prod_stats.iterrows(), start=4):
            ws_prod[f'A{i}'] = i - 3
            ws_prod[f'B{i}'] = str(r['商品详情'])[:60]
            ws_prod[f'C{i}'] = int(r['破损次数'])
            ws_prod[f'D{i}'] = float(r['破损金额'])

        if len(prod_stats) >= 2:
            n = min(15, len(prod_stats))
            chart = BarChart()
            chart.title = "产品破损金额 TOP"
            chart.style = 10
            chart.type = "bar"
            data_ref = Reference(ws_prod, min_col=4, min_row=3, max_row=3+n)
            cat_ref = Reference(ws_prod, min_col=2, min_row=4, max_row=3+n)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cat_ref)
            chart.width = 26
            chart.height = 14
            ws_prod.add_chart(chart, "F3")


    # ========================================================================
    # Sheet 6: 责任方饼图
    # ========================================================================
    ws_pie = wb.create_sheet("责任方占比")
    if '责任方' in df.columns and '金额' in df.columns:
        resp_top = df.groupby('责任方')['金额'].sum().sort_values(ascending=False).head(8)
        ws_pie['A1'] = "责任方破损金额占比 TOP8"
        ws_pie['A1'].font = hdr_font
        ws_pie['A3'] = "责任方"
        ws_pie['B3'] = "破损金额"
        ws_pie['C3'] = "占比"

        total = resp_top.sum()
        for i, (name, amt) in enumerate(resp_top.items(), start=4):
            ws_pie[f'A{i}'] = str(name)
            ws_pie[f'B{i}'] = float(amt)
            ws_pie[f'C{i}'] = float(amt / total) if total > 0 else 0
            ws_pie[f'C{i}'].number_format = '0.0%'

        if len(resp_top) > 0:
            # 使用柱状图代替饼图（openpyxl原生不支持饼图依赖颜色正确显示）
            chart = BarChart()
            chart.type = "col"
            chart.title = "责任方破损金额 TOP8"
            chart.style = 10
            chart.y_axis.title = "金额 (¥)"
            n = len(resp_top)
            data_ref = Reference(ws_pie, min_col=2, min_row=3, max_row=3+n)
            cat_ref = Reference(ws_pie, min_col=1, min_row=4, max_row=3+n)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cat_ref)
            chart.width = 22
            chart.height = 13
            ws_pie.add_chart(chart, "E3")

    # 保存
    wb.save(output_path)



if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
